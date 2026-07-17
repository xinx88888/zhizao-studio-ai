# -*- coding: utf-8 -*-
import os, sys, sqlite3, datetime, json, uuid, urllib.request
from functools import wraps
from flask import Flask, request, session, jsonify, Response
from werkzeug.security import generate_password_hash, check_password_hash
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

# ===== Cloud-ready config =====
DATABASE_URL = os.environ.get("DATABASE_URL", "")  # PostgreSQL URL for cloud, empty = SQLite
FINANCE_EXCEL = os.environ.get("FINANCE_EXCEL", r"E:\Desktop\财务系统-收入支出管理1.xlsx")
SECRET_KEY = os.environ.get("SECRET_KEY", "studio_v4_secret_2024")

# PostgreSQL support (lazy import)
_PG_CONN = None
def _get_pg():
    global _PG_CONN
    if _PG_CONN is None:
        import psycopg2
        _PG_CONN = psycopg2.connect(DATABASE_URL, sslmode="require")
    return _PG_CONN

app = Flask(__name__, static_folder="static")
app.secret_key = SECRET_KEY

@app.after_request
def _no_cache(resp):
    resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    resp.headers["Pragma"] = "no-cache"
    resp.headers["Expires"] = "0"
    return resp

EXCEL_PATH = FINANCE_EXCEL
def _ensure_excel():
    import os.path as _osp
    if not _osp.exists(EXCEL_PATH):
        try:
            from openpyxl import Workbook
            wb = Workbook()
            ws_in = wb.active
            ws_in.title = "收入明细表"
            ws_in.append(["日期", "类别", "项目", "金额", "账号", "备注"])
            ws_ex = wb.create_sheet("支出明细表")
            ws_ex.append(["日期", "类别", "项目", "金额", "账号", "备注"])
            os.makedirs(_osp.dirname(EXCEL_PATH) or ".", exist_ok=True)
            wb.save(EXCEL_PATH)
        except Exception:
            pass
_ensure_excel()

def load_excel():
    _ensure_excel()
    wb = __import__("openpyxl").load_workbook(EXCEL_PATH, data_only=True)
    ti, te, rows = 0, 0, []
    # Read income sheet (收入明细表)
    try:
        ws_in = wb["收入明细表"]
        for i, row in enumerate(ws_in.iter_rows(values_only=True)):
            if i < 2: continue
            if not row[0]: continue
            try:
                date_val = row[0]
                if hasattr(date_val, 'strftime'):
                    date_str = date_val.strftime("%Y-%m-%d")
                else:
                    date_str = str(date_val)
                category = str(row[1] or "")
                project = str(row[2] or "")
                amount = float(row[3] or 0)
                account = str(row[4] or "")
                note = str(row[5] or "")
                if amount > 0:
                    ti += amount
                    rows.append({"date": date_str, "type": category + " - " + project, "amount": amount, "note": note, "who": account})
            except (ValueError, TypeError):
                pass
    except KeyError:
        pass
    # Read expense sheet (支出明细表)
    try:
        ws_ex = wb["支出明细表"]
        for i, row in enumerate(ws_ex.iter_rows(values_only=True)):
            if i < 2: continue
            if not row[0]: continue
            try:
                date_val = row[0]
                if hasattr(date_val, 'strftime'):
                    date_str = date_val.strftime("%Y-%m-%d")
                else:
                    date_str = str(date_val)
                category = str(row[1] or "")
                project = str(row[2] or "")
                amount = float(row[3] or 0)
                account = str(row[4] or "")
                note = str(row[5] or "")
                if amount > 0:
                    te += amount
                    rows.append({"date": date_str, "type": category + " - " + project, "amount": -amount, "note": note, "who": account})
            except (ValueError, TypeError):
                pass
    except KeyError:
        pass
    rows.sort(key=lambda x: x["date"], reverse=True)
    return {"rows": rows, "summary": {"total_income": ti, "total_expense": te, "balance": ti - te}}

def append_excel(income=0, expense=0, note="", who=""):
    _ensure_excel()
    wb = __import__("openpyxl").load_workbook(EXCEL_PATH)
    today = datetime.date.today().strftime("%Y-%m-%d")
    if income > 0:
        try:
            ws = wb["收入明细表"]
            rn = ws.max_row + 1
            ws.cell(row=rn, column=1, value=today)
            ws.cell(row=rn, column=2, value="其他收入")
            ws.cell(row=rn, column=3, value="AI记账")
            ws.cell(row=rn, column=4, value=income)
            ws.cell(row=rn, column=5, value=who)
            ws.cell(row=rn, column=6, value=note)
        except KeyError:
            pass
    if expense > 0:
        try:
            ws = wb["支出明细表"]
            rn = ws.max_row + 1
            ws.cell(row=rn, column=1, value=today)
            ws.cell(row=rn, column=2, value="其他支出")
            ws.cell(row=rn, column=3, value="AI记账")
            ws.cell(row=rn, column=4, value=expense)
            ws.cell(row=rn, column=5, value=who)
            ws.cell(row=rn, column=6, value=note)
        except KeyError:
            pass
    wb.save(EXCEL_PATH)

DB_PATH = os.path.join(os.path.dirname(__file__), "studio.db")
def db_conn():
    if DATABASE_URL:
        return _get_pg()
    return sqlite3.connect(DB_PATH)
def _norm_sql(sql):
    # Convert SQLite placeholders ? to PostgreSQL %s
    if DATABASE_URL:
        return sql.replace("?", "%s")
    return sql
def db_write(sql, args=()):
    conn = db_conn(); cur = conn.cursor()
    cur.execute(_norm_sql(sql), args); conn.commit(); conn.close()
def db_exec(sql, args=()):
    conn = db_conn(); cur = conn.cursor()
    cur.execute(_norm_sql(sql), args); rows = cur.fetchall(); conn.commit(); conn.close()
    return rows

if DATABASE_URL:
    db_write("CREATE TABLE IF NOT EXISTS tasks (id SERIAL PRIMARY KEY, user VARCHAR(50), title TEXT, status VARCHAR(20) DEFAULT 'todo', priority VARCHAR(20) DEFAULT 'medium', due TEXT, created_at TEXT)")
else:
    db_write("CREATE TABLE IF NOT EXISTS tasks (id INTEGER PRIMARY KEY AUTOINCREMENT, user TEXT, title TEXT, status TEXT DEFAULT 'todo', priority TEXT DEFAULT 'medium', due TEXT, created_at TEXT)")
if DATABASE_URL:
    db_write("CREATE TABLE IF NOT EXISTS scripts (id SERIAL PRIMARY KEY, user VARCHAR(50), content_type TEXT, product TEXT, result TEXT, created_at TEXT)")
else:
    db_write("CREATE TABLE IF NOT EXISTS scripts (id INTEGER PRIMARY KEY AUTOINCREMENT, user TEXT, content_type TEXT, product TEXT, result TEXT, created_at TEXT)")
if DATABASE_URL:
    db_write("CREATE TABLE IF NOT EXISTS research (id SERIAL PRIMARY KEY, user VARCHAR(50), research_type TEXT, extra TEXT, result TEXT, created_at TEXT)")
else:
    db_write("CREATE TABLE IF NOT EXISTS research (id INTEGER PRIMARY KEY AUTOINCREMENT, user TEXT, research_type TEXT, extra TEXT, result TEXT, created_at TEXT)")

# ===== Site settings & dynamic team (persisted to JSON) =====
SETTINGS_PATH = os.path.join(os.path.dirname(__file__), "studio_settings.json")
TEAM_EXTRA_PATH = os.path.join(os.path.dirname(__file__), "studio_team_extra.json")
DEFAULT_SETTINGS = {
    "site_title": "智造工作室 AI 工作台",
    "theme": "dark",
    "auto_reload_seconds": 0,
    "live_sync_seconds": 8,
}
def load_settings():
    try:
        with open(SETTINGS_PATH, encoding="utf-8") as f:
            s = json.load(f)
        base = dict(DEFAULT_SETTINGS)
        base.update(s)
        return base
    except Exception:
        return dict(DEFAULT_SETTINGS)
def save_settings(s):
    with open(SETTINGS_PATH, "w", encoding="utf-8") as f:
        json.dump(s, f, ensure_ascii=False, indent=2)
def load_team_extra():
    try:
        with open(TEAM_EXTRA_PATH, encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return []
def save_team_extra(lst):
    with open(TEAM_EXTRA_PATH, "w", encoding="utf-8") as f:
        json.dump(lst, f, ensure_ascii=False, indent=2)

# ===== Users table (DB-backed with hashed passwords, role-based access) =====
if DATABASE_URL:
    db_write("""CREATE TABLE IF NOT EXISTS users (
        id SERIAL PRIMARY KEY,
        username VARCHAR(50) UNIQUE NOT NULL,
        password_hash VARCHAR(255) NOT NULL,
        role VARCHAR(20) NOT NULL DEFAULT 'editor',
        display_name VARCHAR(100),
        skill TEXT,
        description TEXT,
        status VARCHAR(20) NOT NULL DEFAULT 'active',
        created_at TEXT
    )""")
else:
    db_write("""CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT UNIQUE NOT NULL,
        password_hash TEXT NOT NULL,
        role TEXT NOT NULL DEFAULT 'editor',
        display_name TEXT,
        skill TEXT,
        description TEXT,
        status TEXT NOT NULL DEFAULT 'active',
        created_at TEXT
    )""")

def _seed_users_if_empty():
    rows = db_exec("SELECT COUNT(*) FROM users")
    if rows and rows[0][0] == 0:
        import datetime as _dt
        now = _dt.datetime.now().isoformat()
        initial = [
            ("熊科瑞", "智造2026", "admin", "熊科瑞", "建模/3D打印, 短视频拍摄, PPT制作", "团队负责人，统筹管理"),
            ("韦硕",   "智造2026", "editor", "韦硕", "建模/3D打印, 电路焊接",       "建模打印主力"),
            ("邵森",   "智造2026", "editor", "邵森", "PCB设计, 电路焊接",             "电路硬件担当"),
            ("张亦昕", "智造2026", "editor", "张亦昕", "Arduino编程, 电子制作",        "编程主力"),
            ("余晓莉", "智造2026", "editor", "余晓莉", "上色/后处理, 包装发货",         "后处理专家"),
            ("姜润恒", "智造2026", "editor", "姜润恒", "批量生产 (9月加入)",             "9月产能担当"),
        ]
        for u, pwd, role, dn, sk, ds in initial:
            db_write("INSERT INTO users (username, password_hash, role, display_name, skill, description, created_at) VALUES (?,?,?,?,?,?,?)",
                (u, generate_password_hash(pwd), role, dn, sk, ds, now))

_seed_users_if_empty()

def load_user(username):
    rows = db_exec("SELECT id, username, password_hash, role, display_name, skill, description, status, created_at FROM users WHERE username=?", (username,))
    if not rows: return None
    r = rows[0]
    return {
        "id": r[0], "username": r[1], "password_hash": r[2], "role": r[3],
        "display_name": r[4], "skill": r[5], "description": r[6],
        "status": r[7], "created_at": r[8]
    }

def current_user():
    u = session.get("user")
    if not u: return None
    return load_user(u)

def require_role(*allowed):
    from functools import wraps
    def decorator(f):
        @wraps(f)
        def wrapped(*args, **kwargs):
            cu = current_user()
            if not cu: return jsonify({"error": "未登录"}), 401
            if cu["status"] != "active": return jsonify({"error": "账号已禁用"}), 403
            if cu["role"] not in allowed: return jsonify({"error": "权限不足，需要角色: " + "/".join(allowed)}), 403
            return f(*args, **kwargs)
        return wrapped
    return decorator

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user" not in session: return jsonify({"error": "未登录"}), 401
        return f(*args, **kwargs)
    return decorated




@app.route("/api/auth/login", methods=["POST"])
def api_login():
    data = request.get_json(force=True) if request.is_json else request.form.to_dict()
    user = (data.get("user") or data.get("userName") or data.get("username") or "").strip()
    password = (data.get("password") or data.get("pass") or "").strip()
    u = load_user(user)
    if u and u["status"] == "active" and check_password_hash(u["password_hash"], password):
        session["user"] = u["username"]; session.permanent = True
        return jsonify({"ok": True, "user": u["username"], "role": u["role"], "display_name": u.get("display_name") or u["username"]})
    return jsonify({"error": "用户名或密码错误"}), 401

@app.route("/api/auth/logout", methods=["POST"])
def api_logout(): session.clear(); return jsonify({"ok": True})

@app.route("/api/auth/me")
def api_me():
    cu = current_user()
    if cu:
        return jsonify({"logged_in": True, "user": cu["username"], "role": cu["role"], "display_name": cu.get("display_name") or cu["username"], "skill": cu.get("skill", ""), "status": cu["status"]})
    return jsonify({"logged_in": False})

@app.route("/api/stats")
@login_required
def api_stats():
    finance = load_excel()
    tasks = db_exec("SELECT COUNT(*) FROM tasks")
    done = db_exec("SELECT COUNT(*) FROM tasks WHERE status='done'")
    scripts = db_exec("SELECT COUNT(*) FROM scripts")
    return jsonify({"balance": finance["summary"]["balance"], "income": finance["summary"]["total_income"], "expense": finance["summary"]["total_expense"], "tasks": tasks[0][0] if tasks else 0, "done": done[0][0] if done else 0, "scripts": scripts[0][0] if scripts else 0})

@app.route("/api/tasks", methods=["GET"])
@login_required
def api_tasks():
    rows = db_exec("SELECT id, COALESCE(title,description,'无标题'), COALESCE(status,'todo'), COALESCE(priority,'medium'), COALESCE(due,''), created_at FROM tasks ORDER BY id DESC")
    return jsonify([{"id": r[0], "title": r[1], "status": r[2], "priority": r[3], "due": r[4], "created_at": r[5]} for r in rows])

@app.route("/api/tasks/add", methods=["POST"])
@login_required
@require_role("admin", "editor")
def api_tasks_add():
    data = request.get_json(force=True) if request.is_json else request.form.to_dict()
    title = (data.get("title") or "").strip()
    if not title: return jsonify({"error": "标题不能为空"}), 400
    db_write("INSERT INTO tasks (user, title, status, priority, due, created_at) VALUES (?,?,?,?,?,?)",
             (session["user"], title, "todo", data.get("priority","medium"), data.get("due","") or "", datetime.datetime.now().isoformat()))
    return jsonify({"ok": True})

@app.route("/api/tasks/update", methods=["POST"])
@login_required
@require_role("admin", "editor")
def api_tasks_update():
    data = request.get_json(force=True) if request.is_json else request.form.to_dict()
    if data.get("id") and data.get("status"): db_write("UPDATE tasks SET status=? WHERE id=?", (data["status"], data["id"]))
    return jsonify({"ok": True})

@app.route("/api/tasks/delete", methods=["POST"])
@login_required
@require_role("admin", "editor")
def api_tasks_delete():
    data = request.get_json(force=True) if request.is_json else request.form.to_dict()
    if data.get("id"): db_write("DELETE FROM tasks WHERE id=?", (data["id"],))
    return jsonify({"ok": True})

@app.route("/api/team")
@login_required
def api_team():
    rows = db_exec("SELECT username, display_name, skill, description, role FROM users WHERE status='active' ORDER BY id")
    members = [{"name": r[0], "display_name": r[1] or r[0], "skill": r[2] or "", "desc": r[3] or "", "role": r[4]} for r in rows]
    try:
        for m in load_team_extra():
            members.append(m)
    except Exception:
        pass
    return jsonify(members)

@app.route("/api/settings")
@login_required
def api_settings():
    return jsonify(load_settings())

@app.route("/api/settings/update", methods=["POST"])
@login_required
@require_role("admin")
def api_settings_update():
    data = request.get_json(force=True) if request.is_json else request.form.to_dict()
    s = load_settings()
    for k in ["site_title", "theme", "auto_reload_seconds", "live_sync_seconds"]:
        if k in data and data[k] not in (None, ""):
            v = data[k]
            if k in ("auto_reload_seconds", "live_sync_seconds"):
                try:
                    v = int(v)
                except Exception:
                    v = s[k]
            s[k] = v
    save_settings(s)
    return jsonify({"ok": True, "settings": s})



# ===== Admin user management =====
@app.route("/api/admin/users", methods=["GET"])
@login_required
@require_role("admin")
def api_admin_users_list():
    rows = db_exec("SELECT id, username, role, display_name, skill, description, status, created_at FROM users ORDER BY id")
    return jsonify({"ok": True, "users": [
        {"id": r[0], "username": r[1], "role": r[2], "display_name": r[3] or r[1],
         "skill": r[4] or "", "description": r[5] or "", "status": r[6], "created_at": r[7] or ""}
        for r in rows
    ]})

@app.route("/api/admin/users/add", methods=["POST"])
@login_required
@require_role("admin")
def api_admin_users_add():
    data = request.get_json(force=True) if request.is_json else request.form.to_dict()
    username = (data.get("username") or "").strip()
    password = data.get("password") or ""
    role = data.get("role") or "editor"
    display_name = (data.get("display_name") or username).strip()
    skill = (data.get("skill") or "").strip()
    description = (data.get("description") or "").strip()
    if not username or not password:
        return jsonify({"error": "用户名和密码必填"}), 400
    if len(username) < 2 or len(username) > 20:
        return jsonify({"error": "用户名长度 2-20"}), 400
    if len(password) < 4:
        return jsonify({"error": "密码至少 4 位"}), 400
    if role not in ("admin", "editor", "readonly"):
        return jsonify({"error": "无效角色"}), 400
    exists = db_exec("SELECT 1 FROM users WHERE username=?", (username,))
    if exists:
        return jsonify({"error": "用户名已存在"}), 400
    import datetime as _dt
    db_write("INSERT INTO users (username, password_hash, role, display_name, skill, description, created_at) VALUES (?,?,?,?,?,?,?)",
        (username, generate_password_hash(password), role, display_name, skill, description, _dt.datetime.now().isoformat()))
    return jsonify({"ok": True})

@app.route("/api/admin/users/update", methods=["POST"])
@login_required
@require_role("admin")
def api_admin_users_update():
    data = request.get_json(force=True) if request.is_json else request.form.to_dict()
    uid = data.get("id")
    if not uid:
        return jsonify({"error": "缺少用户ID"}), 400
    target = db_exec("SELECT username, role FROM users WHERE id=?", (uid,))
    if not target:
        return jsonify({"error": "用户不存在"}), 404
    updates, params = [], []
    if "role" in data and data["role"] in ("admin", "editor", "readonly"):
        # Don't demote the last admin
        if data["role"] != "admin" and target[0][1] == "admin":
            admins = db_exec("SELECT COUNT(*) FROM users WHERE role='admin' AND status='active'")
            if admins and admins[0][0] <= 1:
                return jsonify({"error": "不能降级最后一个管理员"}), 400
        updates.append("role=?"); params.append(data["role"])
    if "status" in data and data["status"] in ("active", "disabled"):
        # Don't disable the last admin
        if data["status"] == "disabled" and target[0][1] == "admin":
            admins = db_exec("SELECT COUNT(*) FROM users WHERE role='admin' AND status='active'")
            if admins and admins[0][0] <= 1:
                return jsonify({"error": "不能禁用最后一个管理员"}), 400
        updates.append("status=?"); params.append(data["status"])
    if data.get("password"):
        if len(data["password"]) < 4:
            return jsonify({"error": "密码至少 4 位"}), 400
        updates.append("password_hash=?"); params.append(generate_password_hash(data["password"]))
    for f in ("display_name", "skill", "description"):
        if f in data:
            updates.append(f"{f}=?"); params.append(data[f] or "")
    if not updates:
        return jsonify({"error": "没有可更新的字段"}), 400
    params.append(uid)
    db_write(f"UPDATE users SET {','.join(updates)} WHERE id=?", tuple(params))
    return jsonify({"ok": True})

@app.route("/api/admin/users/delete", methods=["POST"])
@login_required
@require_role("admin")
def api_admin_users_delete():
    data = request.get_json(force=True) if request.is_json else request.form.to_dict()
    uid = data.get("id")
    if not uid:
        return jsonify({"error": "缺少用户ID"}), 400
    target = db_exec("SELECT username, role FROM users WHERE id=?", (uid,))
    if not target:
        return jsonify({"error": "用户不存在"}), 404
    if target[0][1] == "admin":
        admins = db_exec("SELECT COUNT(*) FROM users WHERE role='admin' AND status='active'")
        if admins and admins[0][0] <= 1:
            return jsonify({"error": "不能删除最后一个管理员"}), 400
    db_write("DELETE FROM users WHERE id=?", (uid,))
    return jsonify({"ok": True})

@app.route("/api/sync")
@login_required
def api_sync():
    try:
        excel_mtime = int(os.path.getmtime(EXCEL_PATH))
    except Exception:
        excel_mtime = 0
    try:
        tc = db_exec("SELECT COUNT(*), COALESCE(SUM(CASE WHEN status='done' THEN 1 ELSE 0 END),0) FROM tasks")
        task_count = tc[0][0]; done_count = tc[0][1]
    except Exception:
        task_count = 0; done_count = 0
    fin = load_excel()["summary"]
    try:
        tcm = int(os.path.getmtime(os.path.join(os.path.dirname(__file__), "studio_team_extra.json")))
    except Exception:
        tcm = 0
    version = "%s:%s:%s:%s:%s" % (task_count, done_count, round(fin["balance"], 2), excel_mtime, tcm)
    return jsonify({"version": version, "settings": load_settings()})

@app.route("/api/finance")
@login_required
def api_finance(): return jsonify(load_excel())

@app.route("/api/finance/add", methods=["POST"])
@login_required
@require_role("admin", "editor")
def api_finance_add():
    data = request.get_json(force=True) if request.is_json else request.form.to_dict()
    income = float(data.get("income") or 0)
    expense = float(data.get("expense") or 0)
    if income == 0 and expense == 0: return jsonify({"error": "金额不能为0"}), 400
    append_excel(income, expense, data.get("note","") or "", session["user"])
    return jsonify({"ok": True})

@app.route("/api/scripts")
@login_required
def api_scripts():
    rows = db_exec("SELECT id, content_type, product, result, created_at FROM scripts ORDER BY id DESC LIMIT 50")
    return jsonify([{"id": r[0], "content_type": r[1], "product": r[2], "result": r[3], "created_at": r[4]} for r in rows])

@app.route("/api/scripts/add", methods=["POST"])
@login_required
@require_role("admin", "editor")
def api_scripts_add():
    data = request.get_json(force=True) if request.is_json else request.form.to_dict()
    db_write("INSERT INTO scripts (user, content_type, product, result, created_at) VALUES (?,?,?,?,?)",
             (session["user"], data.get("content_type","") or "", data.get("product","") or "", data.get("result","") or "", datetime.datetime.now().isoformat()))
    return jsonify({"ok": True})

@app.route("/api/research")
@login_required
def api_research():
    rows = db_exec("SELECT id, research_type, extra, result, created_at FROM research ORDER BY id DESC LIMIT 50")
    return jsonify([{"id": r[0], "research_type": r[1], "extra": r[2], "result": r[3], "created_at": r[4]} for r in rows])

@app.route("/api/research/add", methods=["POST"])
@login_required
@require_role("admin", "editor")
def api_research_add():
    data = request.get_json(force=True) if request.is_json else request.form.to_dict()
    db_write("INSERT INTO research (user, research_type, extra, result, created_at) VALUES (?,?,?,?,?)",
             (session["user"], data.get("research_type","") or "", data.get("extra","") or "", data.get("result","") or "", datetime.datetime.now().isoformat()))
    return jsonify({"ok": True})

@app.route("/api/ollama_status")
def api_ollama_status():
    try:
        req = urllib.request.Request("http://localhost:11434/api/tags", headers={"Content-Type": "application/json"})
        with urllib.request.urlopen(req, timeout=5) as resp:
            result = json.loads(resp.read())
            return jsonify({"online": True, "models": [m["name"] for m in result.get("models", [])]})
    except:
        return jsonify({"online": False, "models": []})

def ollama_chat(prompt):
    req_data = json.dumps({"model": "qwen2.5:3b", "messages": [{"role": "user", "content": prompt}], "stream": False}).encode()
    req = urllib.request.Request("http://localhost:11434/api/chat", data=req_data, headers={"Content-Type": "application/json"})
    with urllib.request.urlopen(req, timeout=120) as resp:
        return json.loads(resp.read())["message"]["content"]

@app.route("/api/advisor/chat", methods=["POST"])
@login_required
@require_role("admin", "editor")
def api_advisor_chat():
    data = request.get_json(force=True) if request.is_json else request.form.to_dict()
    user_msg = (data.get("message") or "").strip()
    conv = data.get("conversation") or []
    conv.append({"role": "user", "content": user_msg})
    finance = load_excel()
    tasks = db_exec("SELECT COUNT(*) FROM tasks")
    done = db_exec("SELECT COUNT(*) FROM tasks WHERE status='done'")
    tc, dc = (tasks[0][0] if tasks else 0), (done[0][0] if done else 0)
    context = ("你是智造工作室AI战略顾问，6人团队，目标9月新生市场。"
               "当前余额%.2f元，累计支出%.2f元，任务%d/%d条。"
               "请主动提问，帮团队完善9月新生入学布局方案。" % (finance["summary"]["balance"], finance["summary"]["total_expense"], dc, tc))
    try:
        messages = [{"role": "system", "content": context}]
        for m in conv[-10:]: messages.append(m)
        req_data = json.dumps({"model": "qwen2.5:3b", "messages": messages, "stream": False}).encode()
        req = urllib.request.Request("http://localhost:11434/api/chat", data=req_data, headers={"Content-Type": "application/json"})
        with urllib.request.urlopen(req, timeout=90) as resp:
            ai_reply = json.loads(resp.read())["message"]["content"]
    except Exception as e:
        ai_reply = "[AI暂时无法连接：" + str(e) + "]。请检查Ollama是否运行。"
    conv.append({"role": "assistant", "content": ai_reply})
    return jsonify({"reply": ai_reply, "conversation": conv[-12:]})

WIZARD_STEPS = [
    {"id": "target", "question": "第1步：目标定位\n\n你希望智造工作室在新生群体中建立什么样的印象？",
     "options": [
         {"label": "科技感路线", "desc": "主打「用代码造东西」，吸引编程/电子爱好者", "value": "tech"},
         {"label": "创意手作路线", "desc": "主打「亲手做礼物」，吸引文艺/送礼需求", "value": "creative"},
         {"label": "性价比路线", "desc": "主打「学生买得起的黑科技」，吸引价格敏感群体", "value": "budget"},
         {"label": "专业培训路线", "desc": "主打「学3D打印/Arduino」，吸引想学技术的同学", "value": "edu"},
     ]},
    {"id": "products", "question": "第2步：主推产品\n\n针对新生，你最想卖哪类产品？",
     "options": [
         {"label": "发光小夜灯", "desc": "新生必备，宿舍神器，30-60元", "value": "lamp"},
         {"label": "定制游戏手柄", "desc": "个性化外观，50-100元，高利润", "value": "controller"},
         {"label": "手机支架/桌面套装", "desc": "3D打印+Arduino结合，40-80元", "value": "desk"},
         {"label": "毕业礼物定制", "desc": "个性化雕刻，60-150元，高客单价", "value": "gift"},
     ]},
    {"id": "channel", "question": "第3步：推广渠道\n\n你打算主要通过什么方式让新生知道你们？",
     "options": [
         {"label": "抖音/快手短视频", "desc": "制作过程展示视频，引流到微信", "value": "douyin"},
         {"label": "小红书种草", "desc": "图文+视频笔记，吸引女大学生", "value": "xiaohongshu"},
         {"label": "校园摆摊/地推", "desc": "开学季直接接触新生", "value": "campus"},
         {"label": "社团合作", "desc": "赞助社团活动，换取社团群推广", "value": "club"},
     ]},
    {"id": "price", "question": "第4步：定价策略\n\n你的主力产品打算怎么定价？",
     "options": [
         {"label": "低价引流（20-40元）", "desc": "先跑量建口碑，快速起量", "value": "low"},
         {"label": "中等定价（50-80元）", "desc": "平衡利润和销量，推荐", "value": "mid"},
         {"label": "中高定价（80-120元）", "desc": "高品质高价，走利润路线", "value": "high"},
         {"label": "分层定价（20-150全覆盖）", "desc": "引流款+利润款+高端款", "value": "layered"},
     ]},
    {"id": "team_role", "question": "第5步：分工安排\n\n9月新生入学，谁负责什么？",
     "options": [
         {"label": "熊科瑞 — 短视频拍摄剪辑", "desc": "负责抖音/小红书内容产出", "value": "熊科瑞:视频"},
         {"label": "韦硕 — 校园推广/地推", "desc": "开学季摆摊，直接接触新生", "value": "韦硕:地推"},
         {"label": "邵森 — 产品技术支撑", "desc": "确保产品稳定生产和质量", "value": "邵森:技术"},
         {"label": "张亦昕 — 微信客服/社群运营", "desc": "建新生群，维护客户关系", "value": "张亦昕:社群"},
         {"label": "余晓莉 — 包装发货/上色", "desc": "产品包装和快递发货", "value": "余晓莉:发货"},
         {"label": "姜润恒 — 批量产能（9月加入）", "desc": "大批量生产，降成本", "value": "姜润恒:量产"},
     ]},
]

@app.route("/api/wizard/start", methods=["POST"])
@login_required
@require_role("admin", "editor")
def api_wizard_start():
    session["wizard_answers"] = {}
    session["wizard_step"] = 0
    step = WIZARD_STEPS[0]
    return jsonify({"step": 0, "total": len(WIZARD_STEPS), "question": step["question"], "options": step["options"], "progress": 10})

@app.route("/api/wizard/answer", methods=["POST"])
@login_required
@require_role("admin", "editor")
def api_wizard_answer():
    data = request.get_json(force=True) if request.is_json else request.form.to_dict()
    answers = session.get("wizard_answers", {})
    step_idx = session.get("wizard_step", 0)
    answers[WIZARD_STEPS[step_idx]["id"]] = data.get("choice") or data.get("value") or ""
    session["wizard_answers"] = answers
    session["wizard_step"] = step_idx + 1
    if step_idx + 1 >= len(WIZARD_STEPS):
        return jsonify({"done": True, "progress": 100})
    next_step = WIZARD_STEPS[step_idx + 1]
    return jsonify({"step": step_idx + 1, "total": len(WIZARD_STEPS), "question": next_step["question"], "options": next_step["options"], "progress": int((step_idx + 2) / len(WIZARD_STEPS) * 100)})

@app.route("/api/wizard/generate", methods=["POST"])
@login_required
@require_role("admin", "editor")
def api_wizard_generate():
    answers = session.get("wizard_answers", {})
    user = session["user"]
    answers_str = "\n".join([k + ": " + v for k, v in answers.items()])
    prompt = ("你是智造工作室战略规划顾问。基于以下新生布局调研答案，生成完整可执行方案。\n"
              + answers_str + "\n\n"
              + "请输出：1)市场定位 2)产品组合与定价 3)推广时间线 4)各渠道执行计划 5)团队分工表 6)物料清单 7)成本与收益预估 8)风险预案。Markdown格式。")
    try:
        plan = ollama_chat(prompt)
    except Exception as e:
        plan = "[AI生成失败：" + str(e) + "]\n\n已收集答案：\n" + answers_str
    db_write("INSERT INTO research (user, research_type, extra, result, created_at) VALUES (?,?,?,?,?)",
             (user, "wizard", answers_str, plan, datetime.datetime.now().isoformat()))
    session.pop("wizard_answers", None); session.pop("wizard_step", None)
    return jsonify({"plan": plan, "answers": answers})

@app.route("/api/video/generate", methods=["POST"])
@login_required
@require_role("admin", "editor")
def api_video_generate():
    data = request.get_json(force=True) if request.is_json else request.form.to_dict()
    product = data.get("product", "发光小夜灯")
    platform_map = {"douyin": "抖音", "xiaohongshu": "小红书", "bilibili": "B站", "all": "全平台"}
    style_map = {"process": "制作过程展示", "product": "成品效果展示", "story": "创业故事线"}
    platform = platform_map.get(data.get("platform", "douyin"), "抖音")
    style = style_map.get(data.get("style", "process"), "制作过程展示")
    prompt = ("你是短视频策划专家。为智造工作室的" + product + "生成完整短视频方案。\n"
              + "目标平台：" + platform + "，视频风格：" + style + "\n\n"
              + "请生成：1)3条爆款选题（含标题/封面/前3秒钩子）2)完整分镜脚本（30-60秒，标注景别/画面/台词/音效）"
              + "3)拍摄清单 4)剪辑指南（剪映操作+调色+配乐）5)发布时机与标签策略。Markdown格式。")
    try:
        script = ollama_chat(prompt)
    except Exception as e:
        script = "[生成失败：" + str(e) + "]"
    return jsonify({"script": script, "product": product})

@app.route("/api/video/tts", methods=["POST"])
@login_required
@require_role("admin", "editor")
def api_video_tts():
    data = request.get_json(force=True) if request.is_json else request.form.to_dict()
    text = (data.get("text") or "")[:500]
    voice = data.get("voice") or "zh-CN-XiaoxiaoNeural"
    if not text: return jsonify({"error": "文本不能为空"})
    try:
        import subprocess, os
        static_dir = os.path.join(os.path.dirname(__file__), "static")
        os.makedirs(static_dir, exist_ok=True)
        mp3_file = os.path.join(static_dir, "tts_" + uuid.uuid4().hex[:8] + ".mp3")
        r = subprocess.run(["edge-tts", "--text", text, "--voice", voice, "--write-media", mp3_file], capture_output=True, text=True, timeout=30)
        if r.returncode == 0 and os.path.exists(mp3_file) and os.path.getsize(mp3_file) > 1000:
            return jsonify({"ok": True, "audio_url": "/static/" + os.path.basename(mp3_file), "text": text})
        raise Exception("edge-tts failed: " + r.stderr)
    except FileNotFoundError:
        return jsonify({"ok": False, "text": text, "hint": "edge-tts未安装，已返回文本内容。", "fallback_text": True})
    except Exception as e:
        return jsonify({"ok": False, "text": text, "hint": "配音服务异常：" + str(e) + "，已返回文本。", "fallback_text": True})

@app.route("/api/research/web", methods=["POST"])
@login_required
@require_role("admin", "editor")
def api_research_web():
    data = request.get_json(force=True) if request.is_json else request.form.to_dict()
    query = (data.get("query") or "").strip()
    extra_prompt = (data.get("prompt") or data.get("extra") or "").strip()
    user = session["user"]
    if not query: return jsonify({"error": "搜索关键词不能为空"})
    search_results = []
    try:
        # 首选：内置 Bing 搜索（requests 库，无压缩问题）
        from bing_search import bing_search
        search_results = bing_search(query, max_results=10)
    except Exception:
        try:
            # 后备：ddgs 库（并发搜索，可能更稳定但有时超时）
            from ddgs import DDGS
            with DDGS() as ddgs:
                for r in ddgs.text(query, max_results=10):
                    search_results.append({"title": r.get("title",""), "url": r.get("href",""), "snippet": r.get("body","")[:300]})
        except Exception as e2:
            return jsonify({"error": f"搜索失败：{str(e2)}"})
    if not search_results: return jsonify({"error": "未找到相关结果"})
    results_text = "\n".join(["- " + r["title"] + "\n  " + r["snippet"] + "\n  来源：" + r["url"] for r in search_results[:8]])
    analysis_prompt = ("基于以下搜索结果，为智造工作室（3D打印+Arduino，大学生市场）生成分析报告：\n\n"
                      + results_text + "\n\n" + (extra_prompt if extra_prompt else "请生成：市场概况/竞品分析/机会点/建议策略，Markdown格式。"))
    try:
        analysis = ollama_chat(analysis_prompt)
    except Exception as e:
        analysis = "[AI分析失败：" + str(e) + "]"
    full_result = "搜索词：" + query + "\n\n## 搜索结果\n" + results_text + "\n\n## AI分析\n" + analysis
    db_write("INSERT INTO research (user, research_type, extra, result, created_at) VALUES (?,?,?,?,?)",
             (user, "web_search", query, full_result, datetime.datetime.now().isoformat()))
    return jsonify({"ok": True, "results": search_results, "analysis": analysis, "query": query})

RESEARCH_PROMPTS = {
    "competitor": "请分析3D打印+Arduino小夜灯市场的主要竞品，包括价格、功能、销量、用户评价，并指出智造工作室的差异化优势。",
    "trend": "请分析2024-2025年大学生文创/科技类产品的消费趋势，特别是开学季热门品类。",
    "topic": "请为智造工作室推荐10个抖音/小红书爆款选题，包含标题、封面文案、前3秒钩子。",
    "price": "请为3D打印+Arduino小夜灯制定定价策略，分析竞品价格带，给出引流款/利润款/高端款的具体定价建议。",
    "platform": "分析抖音和小红书对3D打印/手工类内容的算法偏好，给出具体的内容策略和标签建议。",
    "full": "请为智造工作室生成一份完整的市场分析报告，包含市场规模、竞品分析、目标用户画像、SWOT分析、4P营销策略。",
}

@app.route("/api/research/ai", methods=["POST"])
@login_required
@require_role("admin", "editor")
def api_research_ai():
    data = request.get_json(force=True) if request.is_json else request.form.to_dict()
    research_type = data.get("type") or "full"
    extra = (data.get("extra") or "").strip()
    user = session["user"]
    prompt = RESEARCH_PROMPTS.get(research_type, RESEARCH_PROMPTS["full"])
    if extra: prompt += "\n\n补充信息：" + extra
    try:
        report = ollama_chat(prompt)
    except Exception as e:
        report = "[AI调研失败：" + str(e) + "]"
    db_write("INSERT INTO research (user, research_type, extra, result, created_at) VALUES (?,?,?,?,?)",
             (user, research_type, extra, report, datetime.datetime.now().isoformat()))
    return jsonify({"report": report})

@app.route("/api/code/arduino", methods=["POST"])
@login_required
@require_role("admin", "editor")
def api_code_arduino():
    data = request.get_json(force=True) if request.is_json else request.form.to_dict()
    task = (data.get("task") or "").strip()
    if not task: return jsonify({"error": "请描述Arduino任务"})
    prompt = ("你是Arduino/单片机开发工程师。请为以下任务生成完整的Arduino代码（含注释和硬件接线说明）。\n\n"
              + "任务：" + task + "\n\n请输出：1)硬件清单（Arduino型号、传感器、接线）2)完整代码（含注释）3)关键函数说明。")
    try:
        code = ollama_chat(prompt)
    except Exception as e:
        code = "[AI代码生成失败：" + str(e) + "]"
    return jsonify({"code": code})









@app.route("/api/command", methods=["POST"])
@login_required
@require_role("admin", "editor")
def api_command():
    data = request.get_json(force=True) if request.is_json else request.form.to_dict()
    cmd = (data.get("command") or data.get("message") or "").strip()
    user = session["user"]
    if not cmd:
        return jsonify({"error": "指令不能为空"})
    
    # Keyword-based action detection (fast + reliable)
    action = "other"
    # Site/System commands (distinct keywords, highest priority)
    if any(w in cmd for w in ["改标题","网站标题","设置标题","改网站标题"]):
        action = "set_title"
    elif any(w in cmd for w in ["切换主题","换主题","改主题","主题切换"]):
        action = "set_theme"
    elif any(w in cmd for w in ["加成员","添加成员","新增成员","加队友"]):
        action = "add_member"
    elif any(w in cmd for w in ["周报","每周汇报","周总结","周汇报","生成周报"]):
        action = "weekly_report"
    # Priority order matters: finance first (most specific), then task_add, task_done, query
    elif any(w in cmd for w in ["记账","收入","支出","报销","花费","付钱","付款"]):
        action = "finance_add"
    elif any(w in cmd for w in ["添加","新建","创建","增加"]) and any(w in cmd for w in ["任务","待办","todo","事项"]):
        action = "task_add"
    elif any(w in cmd for w in ["完成","标记","搞定","done"]) and any(w in cmd for w in ["任务","待办","todo","完成"]):
        action = "task_done"
    elif any(w in cmd for w in ["查询","余额","财务","收支","统计","报表"]):
        action = "finance_query"
    
    # Extract parameters using regex on command
    import re
    title = cmd[max(0, cmd.find("添加")+2 if "添加" in cmd else cmd.find("新建")+2 if "新建" in cmd else 0):][:50].strip()
    tm = re.search(r"(?:添加|新建|创建|任务)[：:]\s*(.{2,50})", cmd)
    if tm: title = tm.group(1).strip()
    
    # Extract amount
    am = re.search(r"(\d+\.?\d*)", cmd)
    amount = float(am.group(1)) if am else 0.0
    # Check if command contains both income AND expense keywords (mixed transaction)
    has_zc_phrase = ("支出" in cmd and ("支出" in cmd.split("收入")[0] if "收入" in cmd else "支出" in cmd))
    has_sr_phrase = "收入" in cmd or "进账" in cmd or "赞助" in cmd or "收款" in cmd
    has_zc_word = "支出" in cmd or "花费" in cmd or "付" in cmd  # word-level check
    is_income = has_sr_phrase and not has_zc_word
    
    # Extract note
    nm = re.search(r"(?:备注|说明|内容)[：:]\s*(.{0,50})", cmd)
    note = nm.group(1).strip() if nm else cmd[:30]
    
    # Extract priority
    priority = "medium"
    if any(w in cmd for w in ["高","紧急","重要","急"]): priority = "high"
    elif any(w in cmd for w in ["低","次要"]): priority = "low"
    
    # Execute action
    finance = load_excel()
    if action == "task_add":
        db_write("INSERT INTO tasks (user, title, status, priority, created_at) VALUES (?,?,?,?,?)",
                 (user, title, "todo", priority, datetime.datetime.now().isoformat()))
        reply_text = "已添加任务：「" + title + "」（" + priority + "）"
    elif action == "task_done":
        try:
            row = db_exec("SELECT id FROM tasks WHERE user=? AND status='todo' ORDER BY id DESC LIMIT 1", (user,))
            if row:
                db_write("UPDATE tasks SET status='done' WHERE id=?", (row[0][0],))
                reply_text = "已将最近一条待办标记为完成"
            else:
                reply_text = "没有待完成的任务可标记"
        except Exception as e:
            reply_text = "标记失败：" + str(e)[:80]
    elif action == "finance_add":
        # Handle mixed transactions (both income and expense)
        if "支出" in cmd and "收入" in cmd:
            amounts = re.findall(r"(\d+\.?\d*)", cmd)
            inc = float(amounts[-1]) if len(amounts) >= 2 else 0.0
            exp = float(amounts[0]) if len(amounts) >= 1 else 0.0
            if inc > 0 or exp > 0:
                append_excel(inc, exp, note, user)
                reply_text = "已记账：收入" + str(inc) + "元 / 支出" + str(exp) + "元（" + note + "）"
            else:
                reply_text = "金额识别失败"
        else:
            inc = amount if is_income else 0.0
            exp = amount if not is_income else 0.0
            if inc > 0 or exp > 0:
                append_excel(inc, exp, note, user)
                reply_text = "已记账：" + ("收入" if is_income else "支出") + " " + str(amount) + "元（" + note + "）"
            else:
                reply_text = "无法识别金额，建议格式：记账收入500备注社团赞助"
    elif action == "finance_query":
        f2 = finance["summary"]
        tc = db_exec("SELECT COUNT(*) FROM tasks")[0][0]
        dc = db_exec("SELECT COUNT(*) FROM tasks WHERE status='done'")[0][0]
        reply_text = "余额" + str(round(f2["balance"], 2)) + "元（收" + str(round(f2["total_income"], 2)) + "元，支" + str(round(f2["total_expense"], 2)) + "元）| 任务" + str(dc) + "/" + str(tc) + "完成"
    elif action == "set_title":
        title = cmd
        for kw in ["改网站标题", "改标题", "网站标题", "设置标题", "标题"]:
            if kw in title:
                title = title[title.find(kw) + len(kw):]
                break
        title = title.lstrip("为：:改成变成是→ ").strip()
        if not title:
            reply_text = "请提供新标题，例如：改标题为智造工作室2.0"
        else:
            s = load_settings(); s["site_title"] = title; save_settings(s)
            reply_text = "已修改网站标题为：「" + title + "」"
    elif action == "set_theme":
        t = "dark"
        if any(w in cmd for w in ["亮","浅","白","light"]):
            t = "light"
        elif any(w in cmd for w in ["绿","青","green","清新"]):
            t = "green"
        else:
            t = "dark"
        s = load_settings(); s["theme"] = t; save_settings(s)
        reply_text = "已切换主题为：" + ("亮色" if t == "light" else "绿色" if t == "green" else "暗色")
    elif action == "add_member":
        txt = cmd
        for kw in ["加成员","添加成员","新增成员","加队友","添加队友"]:
            if kw in txt:
                txt = txt[txt.find(kw) + len(kw):]
                break
        txt = txt.strip()
        name = txt; skill = ""
        for sep in ["技能：","技能:","技能","负责：","负责:","负责","擅长：","擅长:","擅长","：",":"]:
            if sep in txt:
                idx = txt.find(sep); name = txt[:idx].strip(); skill = txt[idx+len(sep):].strip(); break
        if not name:
            reply_text = "请提供成员姓名，例如：加成员 张三 技能：3D打印"
        else:
            extra_list = load_team_extra()
            extra_list.append({"name": name, "skill": skill or "待定", "desc": skill or "新成员", "role": "member"})
            save_team_extra(extra_list)
            reply_text = "已添加成员：「" + name + "」（技能：" + (skill or "待定") + "）"
    elif action == "weekly_report":
        try:
            fin = load_excel()["summary"]
        except Exception:
            fin = {"balance":0,"total_income":0,"total_expense":0}
        try:
            tc = db_exec("SELECT COUNT(*), COALESCE(SUM(CASE WHEN status='done' THEN 1 ELSE 0 END),0) FROM tasks")
            task_count = tc[0][0]; done_count = tc[0][1]
        except Exception:
            task_count = 0; done_count = 0
        try:
            extra_count = len(load_team_extra())
        except Exception:
            extra_count = 0
        prompt = ("你是智造工作室的周报助手。请生成本周工作周报。\n数据：余额%.2f元，累计收入%.2f元，累计支出%.2f元；任务共%d条已完成%d条；团队共%d人。\n请输出Markdown：一、本周数据概览 二、任务进展 三、财务健康度 四、下周行动建议。简洁专业。" % (fin["balance"], fin["total_income"], fin["total_expense"], task_count, done_count, len(USERS)+extra_count))
        try:
            reply_text = ollama_chat(prompt)
        except Exception as e:
            reply_text = "[周报生成失败：" + str(e) + "]"
    else:
        # Fall back to AI for complex commands
        try:
            task_rows = db_exec("SELECT id, COALESCE(title,description,'无标题'), status FROM tasks ORDER BY id DESC LIMIT 5")
        except:
            task_rows = []
        fin = finance["summary"]
        task_list = chr(10).join([str(r[0]) + ". [" + str(r[2]) + "] " + str(r[1]) for r in task_rows]) or "（暂无）"
        prompt = (
            "你是智造工作室助手。用户说：「" + cmd + "」。余额" + str(round(fin["balance"], 2)) +
            "元，任务：" + task_list + "。简洁回复，30字内，像朋友说话。")
        try:
            reply_text = ollama_chat(prompt)[:150]
        except Exception as e:
            reply_text = "收到指令：" + cmd
    
    result = {"ok": True, "reply": reply_text, "action": action}
    if action == "weekly_report":
        result["report"] = reply_text
    if action in ("set_title", "set_theme"):
        result["settings"] = load_settings()
    return jsonify(result)



# ===== HTML =====
html_path = os.path.join(os.path.dirname(__file__), "studio_app_v4.html")
try:
    with open(html_path, encoding="utf-8") as f:
        HTML = f.read()
    HTML_BYTES = HTML.encode("utf-8")
except:
    HTML = "<html><body><h1>HTML文件未找到</h1><p>请确保 studio_app_v4.html 在同一目录下。</p></body></html>"
    HTML_BYTES = HTML.encode("utf-8")

@app.route("/")
def index():
    return Response(HTML_BYTES, mimetype="text/html", headers={"Content-Type": "text/html; charset=utf-8", "Content-Length": str(len(HTML_BYTES))})

if __name__ == "__main__":
    print("Starting Flask on http://0.0.0.0:18080 (本机:127.0.0.1, 局域网:10.186.157.187)")
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 18080)), debug=False)
